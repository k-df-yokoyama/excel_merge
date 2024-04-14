import openpyxl

#ToDo: エラー処理
#ToDo: 単体テスト
#ToDo: 仕様書
#      seq ... Status, Completion data, ..., Comment  <- Target Row Idx/Primary Key Value
#                                             ^- Target Col Idx/Key/Value

class RiskList:
    def __init__(self):
        self.primary_key_title = '学籍番号'

        self.file_name = None
        self.sheet_name = None
        self.min_row = 1
        self.min_col = 1
        self.wb = None
        self.ws = None
        self.header_row = None #ws.iter_rows()で取得したヘッダ行(row)をそのまま保持する
        self.primary_key_value_list = []
        
        self.risk_list = []  #RiskListの1行の各列をdictionaryに格納したデータのlist

    def load_from_excel_book(self, file_name, sheet_name, min_row, min_col):
        # print(file_name)
        self.min_row = min_row
        self.min_col = min_col
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.wb = openpyxl.load_workbook(file_name)
        self.ws = self.wb[sheet_name]

        self.get_risk_list()

    def get_risk_list(self):
        # self.header_row = None

        for row in self.ws.iter_rows(min_row=self.min_row, min_col=self.min_col):
            if row[0].row == self.min_row:
                self.header_row = row
            else:
                one_row_stored_in_dict = {}
                for k, v in zip(self.header_row, row):
                    one_row_stored_in_dict[k.value] = v.value
                self.risk_list.append(one_row_stored_in_dict)

    # 主キーの値のリストを取得
    def get_primary_key_value_list(self):
        self.primary_key_value_list = [d.get(self.primary_key_title) for d in self.risk_list]
        # print(primary_key_value_list)
        return self.primary_key_value_list

    # 指定された主キーの値(target_primary_key_value)に対応する行のリスクデータを辞書型で取得
    def get_dictionary_of_one_risk_data(self, target_primary_key_value):
        for d in self.risk_list:
            if d.get(self.primary_key_title) == target_primary_key_value:
                return d

    # リスクターゲットのキー(primary_key_title)で列名を指定された列の値が、
    # 指定した主キー(target_primary_key_value)の値と同じ行の(1から始まる)インデックスを取得する
    def get_row_index_from_primary_key_value(self, target_primary_key_value):
        i = 0
        for d in self.risk_list:
            i += 1
            if d.get(self.primary_key_title) == target_primary_key_value:
                return i

    # カラム名を指定して(1から始まる)列のインデックスを取得
    def get_col_idx_from_col_key(self, col_key):
        # self.header_row
        idx = 0
        for cell in self.header_row:
            idx += 1
            if cell.value == col_key:
                return idx
        return -1

    # 行を主キーの値(target_primary_key_value)、列を列名(target_col_key)で指定して値を設定する
    def set_risk_data(self, target_primary_key_value, target_col_key, value):
        target_row_idx = target_primary_key_value + self.min_row  # タイトル行があるので -1 は不要
        target_col_idx = self.get_col_idx_from_col_key(target_col_key) + self.min_col - 1
        self.ws.cell(target_row_idx, target_col_idx).value = value

    def save_excel_book(self):
        self.wb.save(self.file_name)

    def print_risk_list(self):
        print(self.risk_list)

    def print_rows(self):
        for row in self.ws.iter_rows(min_row=self.min_row, min_col=self.min_col):
            i = 0
            value = ""
            for col in row:
                i += 1
                if i == 1:
                    value += col.value
                else:
                    value = value + "," + col.value
            print(value)


if __name__ == '__main__':

    risk_list = RiskList()
    risk_list.load_from_excel_book("./school_members1.xlsx", "Sheet1", 2, 2)

    print("- print_risk_list()")
    risk_list.print_risk_list()
    primary_key_value_list = risk_list.get_primary_key_value_list()
    print("- primary_key_value_list")
    print(primary_key_value_list)

    target_col_key = '名前'
    target_row_primary_key_value = "026"
    dict_of_one_risk_data = risk_list.get_dictionary_of_one_risk_data(target_row_primary_key_value)
    print("- print risk data of " + target_row_primary_key_value)
    print("risk_data is " + str(dict_of_one_risk_data))
    if dict_of_one_risk_data is not None:
        row_idx = risk_list.get_row_index_from_primary_key_value(target_row_primary_key_value)
        print(" row_idx is " + str(row_idx))
        print(" name is " + dict_of_one_risk_data.get(target_col_key))

    target_row_primary_key_value = "017"
    dict_of_one_risk_data = risk_list.get_dictionary_of_one_risk_data(target_row_primary_key_value)
    print("- print risk data of " + target_row_primary_key_value)
    print("risk_data is " + str(dict_of_one_risk_data))
    if dict_of_one_risk_data is not None:
        row_idx = risk_list.get_row_index_from_primary_key_value(target_row_primary_key_value)
        print(" row_idx is " + str(row_idx))
        print(" name is " + dict_of_one_risk_data.get(target_col_key))

    target_row_primary_key_value = "000"
    dict_of_one_risk_data = risk_list.get_dictionary_of_one_risk_data(target_row_primary_key_value)
    print("- print risk data of " + target_row_primary_key_value)
    print("risk_data is " + str(dict_of_one_risk_data))
    if dict_of_one_risk_data is not None:
        row_idx = risk_list.get_dictionary_of_one_risk_data_row_idx(target_row_primary_key_value)
        print(" row_idx is " + str(row_idx))
        print(" name is " + dict_of_one_risk_data.get(target_col_key))
